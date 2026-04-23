"""生成生产上线台账 Excel 模版。一次性脚本。"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "生产上线台账-模版.xlsx"

# 样式
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="微软雅黑", size=10)
CAT_FONT = Font(name="微软雅黑", size=11, bold=True, color="1F4E78")

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
CAT_FILL = PatternFill("solid", fgColor="DEEBF7")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def style_title(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = CENTER
    ws.row_dimensions[row].height = 28


def style_category(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = CAT_FONT
    cell.fill = CAT_FILL
    cell.alignment = LEFT
    ws.row_dimensions[row].height = 22


def apply_cell(ws, row, col, value="", font=CELL_FONT, align=LEFT, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = align
    cell.border = BORDER
    if fill:
        cell.fill = fill
    return cell


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------------------------------------------------------------------------
# Sheet 1: 基础信息
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "1-基础信息"
set_widths(ws, [22, 40, 22, 40])

style_title(ws, 1, 4, "生产上线台账 · 基础信息")

info = [
    ("上线日期", "YYYY-MM-DD", "版本标签", "backend: v_._._ / frontend: v_._._"),
    ("计划窗口", "HH:MM – HH:MM", "预计停机", "☐ 无 ☐ 有（分钟）"),
    ("变更类型", "功能/修复/配置/依赖/安全/回滚", "风险等级", "☐ 低 ☐ 中 ☐ 高"),
    ("发布负责人", "", "审批人", ""),
    ("后端负责人", "", "前端负责人", ""),
    ("运维负责人", "", "回滚负责人", ""),
    ("测试负责人", "", "通知对象", "内部群 / 客户 / ___"),
    ("变更摘要", "", "", ""),
    ("业务影响", "", "", ""),
    ("关联需求 / 工单", "", "", ""),
]

row = 3
for a, b, c, d in info:
    apply_cell(ws, row, 1, a, font=Font(name="微软雅黑", size=10, bold=True), fill=CAT_FILL)
    if not c and not d:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        apply_cell(ws, row, 2, b)
    else:
        apply_cell(ws, row, 2, b)
        apply_cell(ws, row, 3, c, font=Font(name="微软雅黑", size=10, bold=True), fill=CAT_FILL)
        apply_cell(ws, row, 4, d)
    ws.row_dimensions[row].height = 24
    row += 1

# ---------------------------------------------------------------------------
# Sheet 2: 变更范围
# ---------------------------------------------------------------------------
ws = wb.create_sheet("2-变更范围")
set_widths(ws, [8, 14, 20, 50, 20, 12])

style_title(ws, 1, 6, "变更范围清单")

headers = ["序号", "模块", "类型", "变更内容 / 说明", "责任人", "是否破坏性"]
for i, h in enumerate(headers, 1):
    apply_cell(ws, 2, i, h).fill = HEADER_FILL
style_header(ws, 2, 6)
ws.row_dimensions[2].height = 24

rows = [
    ("后端 · harness", "代码", "", "", "否"),
    ("后端 · app(Gateway)", "代码", "", "", "否"),
    ("后端 · 中间件", "代码", "", "", "否"),
    ("后端 · 沙箱", "代码", "", "", "否"),
    ("后端 · 子代理", "代码", "", "", "否"),
    ("后端 · MCP", "代码", "", "", "否"),
    ("后端 · 技能", "代码", "", "", "否"),
    ("后端 · 记忆", "代码", "", "", "否"),
    ("后端 · IM 渠道", "代码", "", "", "否"),
    ("后端 · 依赖", "pyproject/uv.lock", "", "", "否"),
    ("前端 · 页面", "代码", "", "", "否"),
    ("前端 · 组件", "代码", "", "", "否"),
    ("前端 · 路由", "代码", "", "", "否"),
    ("前端 · 依赖", "package/pnpm-lock", "", "", "否"),
    ("前端 · 环境变量", "env", "", "", "否"),
    ("配置 · config.yaml", "配置", "", "", "否"),
    ("配置 · config.example.yaml (config_version)", "配置", "", "", "否"),
    ("配置 · extensions_config.json", "配置", "", "", "否"),
    ("基础设施 · Docker/Compose", "基础设施", "", "", "否"),
    ("基础设施 · Nginx", "基础设施", "", "", "否"),
    ("数据 · Postgres schema", "数据", "", "", "否"),
    ("数据 · MinIO bucket", "数据", "", "", "否"),
    ("数据 · 向量库", "数据", "", "", "否"),
]

dv_bool = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
ws.add_data_validation(dv_bool)

for idx, (mod, typ, desc, owner, breaking) in enumerate(rows, 1):
    r = 2 + idx
    apply_cell(ws, r, 1, idx, align=CENTER)
    apply_cell(ws, r, 2, mod)
    apply_cell(ws, r, 3, typ, align=CENTER)
    apply_cell(ws, r, 4, desc)
    apply_cell(ws, r, 5, owner, align=CENTER)
    apply_cell(ws, r, 6, breaking, align=CENTER)
    dv_bool.add(f"F{r}")
    if idx % 2 == 0:
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = ALT_FILL
    ws.row_dimensions[r].height = 22

# ---------------------------------------------------------------------------
# Sheet 3: 上线前检查
# ---------------------------------------------------------------------------
ws = wb.create_sheet("3-上线前检查")
set_widths(ws, [8, 16, 50, 14, 14, 18, 26])

style_title(ws, 1, 7, "上线前检查清单（T-1）")

headers = ["序号", "分类", "检查项", "状态", "责任人", "完成时间", "备注"]
for i, h in enumerate(headers, 1):
    apply_cell(ws, 2, i, h)
style_header(ws, 2, 7)
ws.row_dimensions[2].height = 24

items = [
    ("代码", "代码已合入 main 并打 tag"),
    ("代码", "CHANGELOG / 发布说明已更新"),
    ("后端质量", "make lint 通过"),
    ("后端质量", "make test 通过（含 test_harness_boundary.py）"),
    ("前端质量", "pnpm check 通过（lint + typecheck）"),
    ("前端质量", "pnpm build 成功"),
    ("配置", "config.yaml 已按新版本调整"),
    ("配置", "config.example.yaml → config_version 已 bump（若 schema 变）"),
    ("配置", "extensions_config.json 启用/禁用项已确认"),
    ("配置", ".env 新增变量已在所有环境同步"),
    ("备份", "Postgres 备份完成"),
    ("备份", "MinIO 关键 bucket 快照完成"),
    ("备份", "config.yaml / extensions_config.json 已备份"),
    ("预发环境", "预发冒烟全部通过"),
    ("预发环境", "预发性能 / 压测无回归"),
    ("风险评估", "破坏性变更已与相关方对齐"),
    ("风险评估", "回滚方案已演练 / 确认"),
    ("沟通", "内部群已通知上线时间"),
    ("沟通", "客户 / 业务方已通知（如需）"),
    ("沟通", "值班 / 轮值已安排"),
]

dv_status = DataValidation(type="list", formula1='"待办,进行中,已完成,不适用,阻塞"', allow_blank=True)
ws.add_data_validation(dv_status)

for idx, (cat, item) in enumerate(items, 1):
    r = 2 + idx
    apply_cell(ws, r, 1, idx, align=CENTER)
    apply_cell(ws, r, 2, cat, align=CENTER)
    apply_cell(ws, r, 3, item)
    apply_cell(ws, r, 4, "待办", align=CENTER)
    apply_cell(ws, r, 5, "", align=CENTER)
    apply_cell(ws, r, 6, "", align=CENTER)
    apply_cell(ws, r, 7, "")
    dv_status.add(f"D{r}")
    if idx % 2 == 0:
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = ALT_FILL
    ws.row_dimensions[r].height = 22

# ---------------------------------------------------------------------------
# Sheet 4: 发布步骤
# ---------------------------------------------------------------------------
ws = wb.create_sheet("4-发布步骤")
set_widths(ws, [8, 18, 50, 16, 14, 12, 14, 20])

style_title(ws, 1, 8, "发布操作步骤")

headers = ["序号", "阶段", "操作 / 命令", "预期结果", "责任人", "状态", "实际耗时", "备注"]
for i, h in enumerate(headers, 1):
    apply_cell(ws, 2, i, h)
style_header(ws, 2, 8)
ws.row_dimensions[2].height = 24

steps = [
    ("准备", "git fetch --tags && git checkout <tag>", "HEAD 指向正确 tag"),
    ("准备", "python scripts/check.py", "配置检查通过"),
    ("准备", "scripts/deploy.sh 合并 docker-compose.override.yaml", "合并成功"),
    ("构建", "make up（含镜像构建）", "镜像构建完成"),
    ("后端", "make down && make up", "容器全部 Up"),
    ("后端", "执行数据库迁移（若有）", "迁移成功无报错"),
    ("后端", "检查 Gateway /api/health", "HTTP 200"),
    ("后端", "检查 LangGraph :2024", "返回 ok"),
    ("后端", "检查模型 provider 连通", "调用成功"),
    ("前端", "Nginx :2026 加载最新 bundle", "bundle hash 更新"),
    ("前端", "强刷首页 Console/Network 无红", "无报错"),
    ("配置", "extensions_config.json 热加载生效", "MCP/技能列表正确"),
    ("配置", "技能加载日志无错误", "日志 clean"),
    ("验证", "执行冒烟测试（见 Sheet 5）", "全部通过"),
    ("收尾", "通知相关方发布完成", "已通知"),
    ("收尾", "观测窗口开启 2 小时", "监控面板正常"),
]

for idx, (stage, cmd, expect) in enumerate(steps, 1):
    r = 2 + idx
    apply_cell(ws, r, 1, idx, align=CENTER)
    apply_cell(ws, r, 2, stage, align=CENTER)
    apply_cell(ws, r, 3, cmd)
    apply_cell(ws, r, 4, expect)
    apply_cell(ws, r, 5, "", align=CENTER)
    apply_cell(ws, r, 6, "待办", align=CENTER)
    apply_cell(ws, r, 7, "", align=CENTER)
    apply_cell(ws, r, 8, "")
    dv_status.add(f"F{r}")
    if idx % 2 == 0:
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = ALT_FILL
    ws.row_dimensions[r].height = 22

# ---------------------------------------------------------------------------
# Sheet 5: 冒烟测试
# ---------------------------------------------------------------------------
ws = wb.create_sheet("5-冒烟测试")
set_widths(ws, [8, 14, 50, 14, 14, 24])

style_title(ws, 1, 6, "发布后冒烟测试")

headers = ["序号", "分类", "用例", "结果", "执行人", "截图 / 备注"]
for i, h in enumerate(headers, 1):
    apply_cell(ws, 2, i, h)
style_header(ws, 2, 6)
ws.row_dimensions[2].height = 24

cases = [
    ("后端", "创建新会话并发起普通对话，SSE 流正常"),
    ("后端", "沙箱工具调用（代码执行）"),
    ("后端", "内置工具调用"),
    ("后端", "MCP 工具调用"),
    ("后端", "子代理（subagent）调用"),
    ("后端", "文件上传 / 下载"),
    ("后端", "记忆写入 / 读取"),
    ("后端", "Title 生成 / Summarization"),
    ("前端", "首页加载无 Console/Network 报错"),
    ("前端", "登录 / 鉴权流程"),
    ("前端", "工作台聊天页渲染（消息/工件/todos）"),
    ("前端", "设置页：模型 / MCP / 技能面板读写"),
    ("前端", "移动端样式无回归"),
    ("集成", "IM 渠道消息接收 / 回复"),
    ("集成", "Langfuse 链路可见（若已接入）"),
    ("性能", "首条响应 P95 < 基线"),
    ("性能", "SSE 长连接稳定"),
]

dv_pass = DataValidation(type="list", formula1='"通过,失败,阻塞,跳过"', allow_blank=True)
ws.add_data_validation(dv_pass)

for idx, (cat, case) in enumerate(cases, 1):
    r = 2 + idx
    apply_cell(ws, r, 1, idx, align=CENTER)
    apply_cell(ws, r, 2, cat, align=CENTER)
    apply_cell(ws, r, 3, case)
    apply_cell(ws, r, 4, "", align=CENTER)
    apply_cell(ws, r, 5, "", align=CENTER)
    apply_cell(ws, r, 6, "")
    dv_pass.add(f"D{r}")
    if idx % 2 == 0:
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = ALT_FILL
    ws.row_dimensions[r].height = 22

# ---------------------------------------------------------------------------
# Sheet 6: 观测监控
# ---------------------------------------------------------------------------
ws = wb.create_sheet("6-观测监控")
set_widths(ws, [8, 20, 18, 16, 16, 16, 24])

style_title(ws, 1, 7, "发布后观测（建议盯 2 小时）")

headers = ["序号", "指标", "阈值 / 基线", "T+0", "T+30min", "T+2h", "备注"]
for i, h in enumerate(headers, 1):
    apply_cell(ws, 2, i, h)
style_header(ws, 2, 7)
ws.row_dimensions[2].height = 24

metrics = [
    ("Gateway 错误率", "< 0.5%"),
    ("LangGraph 错误率", "< 0.5%"),
    ("首条响应 P95", "< ___ ms"),
    ("SSE 断连率", "< ___%"),
    ("Gateway 内存", "< ___ MB"),
    ("Gateway CPU", "< ___%"),
    ("LangGraph 内存", "< ___ MB"),
    ("前端 JS 错误", "0"),
    ("日志 ERROR 数", "< ___ /分钟"),
    ("日志 CRITICAL 数", "0"),
    ("模型 provider 失败率", "< 1%"),
    ("数据库连接池", "< 80%"),
]

for idx, (m, threshold) in enumerate(metrics, 1):
    r = 2 + idx
    apply_cell(ws, r, 1, idx, align=CENTER)
    apply_cell(ws, r, 2, m)
    apply_cell(ws, r, 3, threshold, align=CENTER)
    for c in (4, 5, 6):
        apply_cell(ws, r, c, "", align=CENTER)
    apply_cell(ws, r, 7, "")
    if idx % 2 == 0:
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = ALT_FILL
    ws.row_dimensions[r].height = 22

# ---------------------------------------------------------------------------
# Sheet 7: 回滚预案
# ---------------------------------------------------------------------------
ws = wb.create_sheet("7-回滚预案")
set_widths(ws, [8, 22, 50, 16, 14, 20])

style_title(ws, 1, 6, "回滚预案")

apply_cell(ws, 2, 1, "触发条件", font=Font(name="微软雅黑", size=10, bold=True), fill=CAT_FILL)
ws.merge_cells("B2:F2")
apply_cell(ws, 2, 2, "错误率 > ___%  或  核心路径不可用持续 > ___ 分钟  或  业务方明确要求")
ws.row_dimensions[2].height = 26

headers = ["序号", "步骤", "操作 / 命令", "责任人", "状态", "备注"]
for i, h in enumerate(headers, 1):
    apply_cell(ws, 4, i, h)
style_header(ws, 4, 6)
ws.row_dimensions[4].height = 24

rollback = [
    ("决策", "确认触发条件并通知负责人"),
    ("代码", "git checkout <上一 tag>"),
    ("部署", "make down && make up"),
    ("配置", "恢复 config.yaml 备份"),
    ("配置", "恢复 extensions_config.json 备份"),
    ("数据", "回滚数据库迁移（若有）"),
    ("数据", "MinIO / 向量库恢复（若有）"),
    ("验证", "重复执行冒烟测试（Sheet 5）"),
    ("通知", "通知相关方回滚完成及影响范围"),
    ("复盘", "安排复盘会（48 小时内）"),
]

for idx, (step, cmd) in enumerate(rollback, 1):
    r = 4 + idx
    apply_cell(ws, r, 1, idx, align=CENTER)
    apply_cell(ws, r, 2, step, align=CENTER)
    apply_cell(ws, r, 3, cmd)
    apply_cell(ws, r, 4, "", align=CENTER)
    apply_cell(ws, r, 5, "待办", align=CENTER)
    apply_cell(ws, r, 6, "")
    dv_status.add(f"E{r}")
    if idx % 2 == 0:
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = ALT_FILL
    ws.row_dimensions[r].height = 22

# ---------------------------------------------------------------------------
# Sheet 8: 发布时间线
# ---------------------------------------------------------------------------
ws = wb.create_sheet("8-发布时间线")
set_widths(ws, [8, 18, 28, 40, 14, 24])

style_title(ws, 1, 6, "发布时间线记录")

headers = ["序号", "时间", "事件 / 阶段", "详情", "执行人", "结果"]
for i, h in enumerate(headers, 1):
    apply_cell(ws, 2, i, h)
style_header(ws, 2, 6)
ws.row_dimensions[2].height = 24

events = [
    "发布窗口开始",
    "代码 checkout 完成",
    "镜像构建完成",
    "后端重启完成",
    "健康检查通过",
    "前端就绪",
    "配置热加载生效",
    "冒烟测试通过",
    "监控窗口开启",
    "发布完成通知",
    "观测窗口结束",
]

for idx, ev in enumerate(events, 1):
    r = 2 + idx
    apply_cell(ws, r, 1, idx, align=CENTER)
    apply_cell(ws, r, 2, "", align=CENTER)
    apply_cell(ws, r, 3, ev)
    apply_cell(ws, r, 4, "")
    apply_cell(ws, r, 5, "", align=CENTER)
    apply_cell(ws, r, 6, "", align=CENTER)
    if idx % 2 == 0:
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = ALT_FILL
    ws.row_dimensions[r].height = 22

# ---------------------------------------------------------------------------
# Sheet 9: 遗留与签字
# ---------------------------------------------------------------------------
ws = wb.create_sheet("9-遗留与签字")
set_widths(ws, [8, 40, 18, 14, 24])

style_title(ws, 1, 5, "遗留问题 / 签字确认")

style_category(ws, 2, 5, "一、遗留问题 / 后续跟进")
headers = ["序号", "问题描述", "跟进负责人", "计划完成", "备注"]
for i, h in enumerate(headers, 1):
    apply_cell(ws, 3, i, h)
style_header(ws, 3, 5)
ws.row_dimensions[3].height = 24

for i in range(1, 9):
    r = 3 + i
    apply_cell(ws, r, 1, i, align=CENTER)
    for c in range(2, 6):
        apply_cell(ws, r, c, "")
    ws.row_dimensions[r].height = 22

style_category(ws, 13, 5, "二、签字确认")
sign_headers = ["角色", "姓名", "签字", "日期", "备注"]
for i, h in enumerate(sign_headers, 1):
    apply_cell(ws, 14, i, h)
style_header(ws, 14, 5)
ws.row_dimensions[14].height = 24

roles = [
    "发布负责人",
    "后端负责人",
    "前端负责人",
    "运维负责人",
    "测试负责人",
    "产品 / 业务负责人",
    "审批人",
]

for i, role in enumerate(roles, 1):
    r = 14 + i
    apply_cell(ws, r, 1, role, align=CENTER, fill=ALT_FILL)
    for c in range(2, 6):
        apply_cell(ws, r, c, "")
    ws.row_dimensions[r].height = 28

# 冻结首行
for sheet in wb.sheetnames:
    wb[sheet].freeze_panes = "A3"

wb.save(OUT)
print(f"✓ 已生成：{OUT}")
